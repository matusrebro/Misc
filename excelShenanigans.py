"""

These are just my notes regarding some of the python
excel activities

"""

import openpyxl # main module for wokring with excel files
from openpyxl.styles import PatternFill # this is used for coloring of cells
from openpyxl.comments import Comment # this is used for adding of the comments

filepath = r'excelFile.xlsx'

# open excel file as a workbook
workbook = openpyxl.load_workbook(filepath)
# load the specific sheet
ws = workbook['Sheet1']

# setting up the red coloring to color the cells with
redFill = PatternFill(start_color='FF0000',
                    end_color='FF0000',
                    fill_type='solid')

# selecting the range for coloring
cell_range  = ws['A1':'A20']

# looping through the range and coloring with color
# that was setuped earlier
for cell in cell_range:
    cell[0].fill = redFill

# comment string
commentString= 'Comment'

# writing a comment to specific cell
ws["C6"].comment = Comment(commentString, 'Steve')

# saving workbook to the path of a original file
workbook.save(filepath)
